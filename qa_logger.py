from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

EXCEL_FILE = "QA_Report.xlsx"

def init_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Execution Log"
        headers = [
            "Timestamp",
            "Test Case ID",
            "Actual Result",
            "Expected Result",
            "Status",
            "Execution Time (s)",
            "Notes"
        ]
        ws.append(headers)
        wb.save(EXCEL_FILE)

def log_result(test_case_id, actual_result, expected_result, status, duration, notes=""):
    init_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Execution Log"]
    new_row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        test_case_id,
        actual_result,
        expected_result,
        status,
        round(duration, 2),
        notes
    ]
    ws.append(new_row)
    wb.save(EXCEL_FILE)
